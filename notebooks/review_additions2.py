"""Extract the uncertain cells of concept_additions_2.xlsx into a review sheet
for a second opinion (ChatGPT), mirroring the earlier welsh_regen / held_rows
delegation format.

A cell is "uncertain" when it is either
  * QC-flagged  (disagrees with Wiktionary and/or NorthEuraLex), or
  * unreferenced (neither reference has that concept x language, so the machine
    translation was never checked against anything).

Verified columns (English/Chinese/French/Irish) are skipped. Each row carries
sense anchors (Chinese, French, Concepticon gloss, German/Russian attestations)
plus blank `proposal` / `note` columns for the reviewer.

    python notebooks/review_additions2.py
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from nguasach.lexibank_qc import _attestations, _agree

SHEET = Path("data/interim/concept_additions_2.xlsx")
WCACHE = Path("data/interim/wiktionary_cache_additions2.json")
VERIFIED = {"English", "Chinese", "French", "Irish"}
# languages where the fill was genuine machine translation (no good attestation
# source) -> reviewer's time is best spent here
HIGH = {"Vietnamese", "Thai", "Indonesian", "Swahili", "Japanese", "Korean",
        "Turkish", "Hebrew", "Hindi"}


def main() -> None:
    ws = openpyxl.load_workbook(SHEET).active
    hdr = [c.value for c in ws[1]]
    H = {n: i for i, n in enumerate(hdr)}
    langs = [h for h in hdr if not h.startswith("_") and h not in VERIFIED]

    wik = json.loads(WCACHE.read_text(encoding="utf-8")) if WCACHE.exists() else {}
    nel = _attestations(Path("data/raw/northeuralex-cldf"))

    out = openpyxl.Workbook()
    o = out.active
    o.title = "review"
    o.append(["tier", "english", "concepticon", "chinese", "french", "irish",
              "de/ru/es anchor", "language", "current", "status",
              "wiktionary_ref", "northeuralex_ref", "proposal", "note"])

    n_flag = n_noref = 0
    by_lang: dict[str, int] = {}
    by_tier: dict[int, int] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        eng = str(r[H["English"]]).strip()
        cid = str(r[H["_concepticon_id"]] or "").strip()
        cg = str(r[H["_concepticon_gloss"]] or "")
        anchor = " / ".join(f"{k}={r[H[k]]}" for k in ("German", "Russian", "Spanish")
                            if r[H[k]])
        wrefs = wik.get(eng, {})
        for L in langs:
            cell = str(r[H[L]] or "").strip()
            if not cell:                              # blanked (e.g. wrong homograph)
                by_lang[L] = by_lang.get(L, 0) + 1
                by_tier[1] = by_tier.get(1, 0) + 1
                o.append([1, eng, f"{cid} {cg}".strip(), r[H["Chinese"]],
                          r[H["French"]], r[H["Irish"]], anchor, L, "",
                          "MISSING", "", "", "", ""])
                n_noref += 1
                continue
            w_forms = set(wrefs.get(L, []))
            n_forms = set(nel.get((cid, L), set())) if cid else set()
            w_ok = _agree(cell, w_forms, L) if w_forms else None
            n_ok = _agree(cell, n_forms, L) if n_forms else None

            if w_ok is False and n_ok is False:
                status, tier = "FLAG:both", 1
                n_flag += 1
            elif n_ok is False:
                status, tier = "FLAG:nel", 1        # NEL is the curated source
                n_flag += 1
            elif w_ok is None and n_ok is None:
                status = "NO-REF"
                tier = 2 if L in HIGH else 4        # 2 = real MT, 4 = attested-ish
                n_noref += 1
            elif w_ok is False:
                status, tier = "FLAG:wik", 3        # Wiktionary multi-sense noise
                n_flag += 1
            else:
                continue                            # reference-confirmed

            by_lang[L] = by_lang.get(L, 0) + 1
            by_tier[tier] = by_tier.get(tier, 0) + 1
            o.append([tier, eng, f"{cid} {cg}".strip(), r[H["Chinese"]], r[H["French"]],
                      r[H["Irish"]], anchor, L, cell, status,
                      " / ".join(sorted(w_forms)[:5]),
                      " / ".join(sorted(n_forms)[:5]), "", ""])

    rows = list(o.iter_rows(min_row=2, values_only=True))
    rows.sort(key=lambda x: (x[0], x[7], x[1]))          # tier, language, english
    for i in range(o.max_row, 1, -1):
        o.delete_rows(i)
    for row in rows:
        o.append(row)

    dst = Path("data/interim/additions2_review.xlsx")
    out.save(dst)
    print(f"wrote {dst}")
    print(f"  {len(rows)} cells  ({n_flag} flagged, {n_noref} unreferenced)")
    _T = {1: "FLAG both / NEL  (real disagreement)",
          2: "NO-REF, MT language  (native check)",
          3: "FLAG wik only  (likely Wiktionary noise, quick scan)",
          4: "NO-REF, attested-ish European  (optional)"}
    for t in (1, 2, 3, 4):
        print(f"  tier {t}: {by_tier.get(t, 0):3}  {_T[t]}")
    print("  by language:")
    for L, c in sorted(by_lang.items(), key=lambda kv: -kv[1]):
        print(f"    {L:12} {c}")


if __name__ == "__main__":
    main()
