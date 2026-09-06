"""Fill the 18 non-verified columns of concept_additions_2.xlsx.

Pass 1: English Wiktionary ``Translations`` sections (reuses wiktionary_qc.fetch).
Pass 2: Google Translate (deep-translator) for whatever Wiktionary left blank.

Verified columns (English / Chinese / French / Irish) and any cell already
filled are never touched. Writes a ``_fill_source`` note column.

    python notebooks/fill_additions2.py
"""
from __future__ import annotations

import time
from pathlib import Path

import openpyxl

from nguasach.wiktionary_qc import fetch

SHEET = Path("data/interim/concept_additions_2.xlsx")
CACHE = Path("data/interim/wiktionary_cache_additions2.json")
VERIFIED = {"English", "Chinese", "French", "Irish"}

# our column -> deep-translator (ISO-639-1) code
GT_CODE = {
    "Hungarian": "hu", "Finnish": "fi", "Greek": "el", "Russian": "ru",
    "German": "de", "Spanish": "es", "Italian": "it", "Welsh": "cy",
    "Vietnamese": "vi", "Japanese": "ja", "Korean": "ko", "Thai": "th",
    "Indonesian": "id", "Turkish": "tr", "Arabic": "ar", "Hebrew": "iw",
    "Swahili": "sw", "Hindi": "hi",
}


def _pick(forms: list[str]) -> str:
    """First Wiktionary form, preferring the shortest single-token one."""
    if not forms:
        return ""
    singles = [f for f in forms if " " not in f] or forms
    return min(singles, key=len)


def main() -> None:
    wb = openpyxl.load_workbook(SHEET)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    H = {n: i + 1 for i, n in enumerate(hdr)}
    cols = [c for c in GT_CODE if c in H]
    rows = list(range(2, ws.max_row + 1))
    words = [str(ws.cell(r, H["English"]).value).strip() for r in rows]

    # ---- pass 1: Wiktionary ----
    print(f"[fill] Wiktionary fetch for {len(words)} words ...")
    wik = fetch(words, CACHE)

    if "_fill_source" not in H:
        ws.cell(1, ws.max_column + 1).value = "_fill_source"
        hdr = [c.value for c in ws[1]]
        H = {n: i + 1 for i, n in enumerate(hdr)}

    filled_wik = {c: 0 for c in cols}
    still_empty: list[tuple[int, str, str]] = []          # (row, col, english)
    for r, w in zip(rows, words):
        refs = wik.get(w, {})
        notes = []
        for c in cols:
            cell = ws.cell(r, H[c])
            if str(cell.value or "").strip():
                continue
            form = _pick(refs.get(c, []))
            if form:
                cell.value = form
                filled_wik[c] += 1
                notes.append(c)
            else:
                still_empty.append((r, c, w))
        if notes:
            prev = str(ws.cell(r, H["_fill_source"]).value or "")
            ws.cell(r, H["_fill_source"]).value = (prev + " wik:" + ",".join(notes)).strip()
    wb.save(SHEET)
    print("[fill] Wiktionary filled per column:",
          {k: v for k, v in filled_wik.items() if v})
    print(f"[fill] {len(still_empty)} cells remain -> Google Translate")

    # ---- pass 2: Google Translate fallback ----
    try:
        from deep_translator import GoogleTranslator
    except Exception as e:                                 # pragma: no cover
        print(f"[fill] deep-translator unavailable ({e}); skipping pass 2")
        _report(ws, H, cols, rows)
        return

    translators: dict[str, object] = {}
    filled_gt = {c: 0 for c in cols}
    fails = {c: 0 for c in cols}
    for i, (r, c, w) in enumerate(still_empty):
        if fails[c] >= 8:                                  # give up on a dead target
            continue
        tr = translators.get(c)
        if tr is None:
            tr = translators[c] = GoogleTranslator(source="en", target=GT_CODE[c])
        try:
            out = (tr.translate(w) or "").strip()
        except Exception:
            fails[c] += 1
            time.sleep(1.0)
            continue
        if out and out.lower() != w.lower():
            ws.cell(r, H[c]).value = out
            filled_gt[c] += 1
            prev = str(ws.cell(r, H["_fill_source"]).value or "")
            if "gt:" not in prev:
                ws.cell(r, H["_fill_source"]).value = (prev + " gt").strip()
        if i % 50 == 0:
            wb.save(SHEET)
            print(f"[fill] GT {i}/{len(still_empty)}")
        time.sleep(0.4)
    wb.save(SHEET)
    print("[fill] Google filled per column:", {k: v for k, v in filled_gt.items() if v})
    _report(ws, H, cols, rows)


def _report(ws, H, cols, rows) -> None:
    print(f"\n{'column':12} filled/total")
    for c in cols:
        n = sum(1 for r in rows if str(ws.cell(r, H[c]).value or "").strip())
        print(f"  {c:12} {n:3}/{len(rows)}   {'' if n == len(rows) else 'MISSING ' + str(len(rows) - n)}")


if __name__ == "__main__":
    main()
