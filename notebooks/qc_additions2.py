"""QC the filled concept_additions_2.xlsx against Wiktionary + NorthEuraLex.

Same logic as the wiktionary-qc stage's ``translation_qc_combined``: a cell is
flagged ``differ_both`` when it disagrees with BOTH references (the strong error
signal), ``differ_wik_only`` / ``differ_nel_only`` for a single-source
disagreement. Verified columns are skipped.

Writes data/interim/additions2_qc.csv  ->  columns the user should eyeball.

    python notebooks/qc_additions2.py
"""
from __future__ import annotations

import csv
import json
from pathlib import Path

import openpyxl

from nguasach.lexibank_qc import _attestations, _agree

SHEET = Path("data/interim/concept_additions_2.xlsx")
WCACHE = Path("data/interim/wiktionary_cache_additions2.json")
VERIFIED = {"English", "Chinese", "French", "Irish"}


def main() -> None:
    ws = openpyxl.load_workbook(SHEET).active
    hdr = [c.value for c in ws[1]]
    H = {n: i for i, n in enumerate(hdr)}
    langs = [h for h in hdr if not h.startswith("_") and h not in VERIFIED]

    wik = json.loads(WCACHE.read_text(encoding="utf-8")) if WCACHE.exists() else {}
    nel = _attestations(Path("data/raw/northeuralex-cldf"))   # {(cid, ourlang): {forms}}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    out = []
    counts = {"differ_both": 0, "differ_wik_only": 0, "differ_nel_only": 0}
    for r in rows:
        eng = str(r[H["English"]]).strip()
        cid = str(r[H["_concepticon_id"]] or "").strip()
        wrefs = wik.get(eng, {})
        for L in langs:
            cell = str(r[H[L]] or "").strip()
            if not cell:
                continue
            w_forms = set(wrefs.get(L, []))
            n_forms = set(nel.get((cid, L), set())) if cid else set()
            w_ok = _agree(cell, w_forms, L) if w_forms else None
            n_ok = _agree(cell, n_forms, L) if n_forms else None
            if w_ok is False and n_ok is False:
                verdict = "differ_both"
            elif w_ok is False and n_ok is None:
                verdict = "differ_wik_only"
            elif n_ok is False and w_ok is None:
                verdict = "differ_nel_only"
            else:
                continue
            counts[verdict] += 1
            out.append({
                "english": eng, "language": L, "ours": cell, "verdict": verdict,
                "wiktionary": " / ".join(sorted(w_forms)[:4]),
                "northeuralex": " / ".join(sorted(n_forms)[:4]),
                "fill_source": str(r[H.get("_fill_source", -1)] or ""),
            })

    out.sort(key=lambda d: (d["verdict"] != "differ_both", d["language"], d["english"]))
    dst = Path("data/interim/additions2_qc.csv")
    with dst.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=["english", "language", "ours", "verdict",
                                           "wiktionary", "northeuralex", "fill_source"])
        w.writeheader()
        w.writerows(out)
    print(f"wrote {dst}  ({len(out)} flags)")
    print("  by verdict:", counts)
    print("\n  differ_both (disagree with both references):")
    for d in out:
        if d["verdict"] == "differ_both":
            print(f"    {d['english']:12} {d['language']:11} ours={d['ours']!r}  "
                  f"wik={d['wiktionary']!r}  nel={d['northeuralex']!r}")


if __name__ == "__main__":
    main()
